import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import re
import io
import subprocess
import sys

def instalar(paquete):
    subprocess.check_call([sys.executable, "-m", "pip", "install", paquete, "-q"])

# Sin langdetect - solo diccionario técnico

# ── Diccionario técnico de repuestos maquinaria pesada ────
DICCIONARIO_TECNICO = {
    # ── Filtros ──
    "filter": "filtro", "filters": "filtros",
    "strainer": "colador", "element": "elemento filtrante",
    "cartridge": "cartucho", "screen": "malla",

    # ── Motor ──
    "engine": "motor", "engines": "motores",
    "piston": "pistón", "pistons": "pistones",
    "cylinder": "cilindro", "cylinders": "cilindros",
    "crankshaft": "cigüeñal", "camshaft": "árbol de levas",
    "conrod": "biela", "liner": "camisa", "liners": "camisas",
    "head": "culata", "heads": "culatas",
    "valve": "válvula", "valves": "válvulas",
    "rocker": "balancín", "rockers": "balancines",
    "lifter": "levantaválvulas", "tappet": "taqué",
    "injector": "inyector", "injectors": "inyectores",
    "nozzle": "tobera", "nozzles": "toberas",
    "turbocharger": "turbo cargador",
    "intake": "admisión", "exhaust": "escape",
    "manifold": "colector", "muffler": "silenciador",
    "starter": "motor de arranque",
    "alternator": "alternador",
    "flywheel": "volante motor",

    # ── Sellos y juntas ──
    "seal": "sello", "seals": "sellos",
    "gasket": "junta", "gaskets": "juntas",
    "o-ring": "junta", "oring": "junta",
    "o´ring": "junta", "o'ring": "junta",
    "retainer": "retén", "retainers": "retenes",
    "wiper": "rascador", "wipers": "rascadores",
    "lip seal": "sello de labio",
    "backup ring": "anillo de respaldo",

    # ── Rodamientos y bujes ──
    "bearing": "rodamiento", "bearings": "rodamientos",
    "bushing": "buje", "bushings": "bujes",
    "sleeve": "manguito", "sleeves": "manguitos",
    "race": "pista", "cone": "cono", "cup": "copa",
    "roller": "rodillo", "rollers": "rodillos",
    "needle": "aguja",

    # ── Transmisión y tren de potencia ──
    "transmission": "transmisión",
    "differential": "diferencial",
    "gearbox": "caja de engranajes",
    "gear": "engranaje", "gears": "engranajes",
    "shaft": "eje", "shafts": "ejes",
    "axle": "eje", "axles": "ejes",
    "coupling": "acople", "couplings": "acoples",
    "clutch": "embrague", "clutches": "embragues",
    "disc": "disco", "discs": "discos",
    "brake": "freno", "brakes": "frenos",
    "sprocket": "rueda dentada", "sprockets": "ruedas dentadas",
    "chain": "cadena", "chains": "cadenas",
    "belt": "correa", "belts": "correas",
    "pulley": "polea", "pulleys": "poleas",
    "idler": "rueda guía",
    "planetary": "planetario",
    "carrier": "portasatélites",
    "sun gear": "engranaje solar",

    # ── Sistema hidráulico ──
    "pump": "bomba", "pumps": "bombas",
    "motor": "motor hidráulico",
    "cylinder": "cilindro", "cylinders": "cilindros",
    "hose": "manguera", "hoses": "mangueras",
    "tube": "tubo", "tubes": "tubos",
    "pipe": "caño", "pipes": "caños",
    "fitting": "fitting", "fittings": "fittings",
    "elbow": "codo", "elbows": "codos",
    "nipple": "niple", "nipples": "niples",
    "flange": "brida", "flanges": "bridas",
    "tee": "tee",
    "adapter": "adaptador", "adapters": "adaptadores",
    "connector": "conector", "connectors": "conectores",
    "relief valve": "válvula de alivio",
    "check valve": "válvula de retención",
    "control valve": "válvula de control",
    "accumulator": "acumulador",
    "reservoir": "depósito",

    # ── Estructura y carrocería ──
    "bracket": "soporte", "brackets": "soportes",
    "plate": "placa", "plates": "placas",
    "cover": "tapa", "covers": "tapas",
    "cap": "tapa", "caps": "tapas",
    "housing": "carcasa", "housings": "carcasas",
    "frame": "bastidor", "frames": "bastidores",
    "guard": "protector", "guards": "protectores",
    "shield": "escudo", "shields": "escudos",
    "panel": "panel", "panels": "paneles",
    "door": "puerta", "doors": "puertas",
    "window": "ventana", "windows": "ventanas",
    "glass": "vidrio",
    "mirror": "espejo", "mirrors": "espejos",

    # ── Fijación y sujeción ──
    "bolt": "perno", "bolts": "pernos",
    "nut": "tuerca", "nuts": "tuercas",
    "washer": "arandela", "washers": "arandelas",
    "screw": "tornillo", "screws": "tornillos",
    "pin": "pasador", "pins": "pasadores",
    "clip": "grapa", "clips": "grapas",
    "ring": "aro", "rings": "aros",
    "snap ring": "anillo elástico",
    "lock": "traba", "locking": "de traba",
    "stud": "espárrago", "studs": "espárragos",

    # ── Suspensión y dirección ──
    "spring": "resorte", "springs": "resortes",
    "shock": "amortiguador", "damper": "amortiguador",
    "tie rod": "barra de dirección",
    "steering": "dirección",
    "knuckle": "mangueta",
    "link": "eslabón", "links": "eslabones",
    "arm": "brazo", "arms": "brazos",
    "rod": "vástago", "rods": "vástagos",

    # ── Tren de rodaje (orugas) ──
    "track": "oruga", "tracks": "orugas",
    "shoe": "zapata", "shoes": "zapatas",
    "pad": "zapata", "pads": "zapatas",
    "link": "eslabón",
    "idler": "rueda guía",
    "roller": "rodillo",
    "blade": "cuchilla", "blades": "cuchillas",
    "bucket": "balde", "buckets": "baldes",
    "boom": "pluma", "booms": "plumas",
    "stick": "brazo", "dipper": "brazo",

    # ── Sistema eléctrico ──
    "switch": "interruptor", "switches": "interruptores",
    "relay": "relé", "relays": "relés",
    "fuse": "fusible", "fuses": "fusibles",
    "sensor": "sensor", "sensors": "sensores",
    "harness": "mazo de cables",
    "wire": "cable", "wires": "cables",
    "lamp": "lámpara", "lamps": "lámparas",
    "gauge": "indicador", "gauges": "indicadores",
    "lever": "palanca", "levers": "palancas",
    "knob": "perilla", "knobs": "perillas",
    "handle": "manija", "handles": "manijas",
    "seat": "asiento", "seats": "asientos",

    # ── Enfriamiento ──
    "radiator": "radiador",
    "fan": "ventilador", "fans": "ventiladores",
    "thermostat": "termostato",
    "cooler": "enfriador", "coolers": "enfriadores",
    "condenser": "condensador",
    "compressor": "compresor",

    # ── Varios ──
    "assembly": "conjunto",
    "kit": "kit",
    "set": "juego",
    "group": "grupo", "gp": "grupo",
    "gp-pr": "grupo de presión",
    "seal_exhaust": "sello de escape",
    "allem": "allen",
    "vibratory": "vibratorio",
    "smooth": "liso",
    "asphalt": "asfalto",
    "single": "simple", "double": "doble",
    "drum": "tambor", "drums": "tambores",
    "pulley": "polea",
    "harness": "mazo de cables",
    "flange": "brida",
}


# ── Lista de palabras españolas para segmentación ─────────
PALABRAS_SEPARACION = set([
    # Generado desde 15.732 artículos del dominio
    "conectorroscado",
    "cabezahexagonal",
    "arandeladeacero",
    "valvuladeescape",
    "discodefriccion",
    "sensordepresion",
    "motordearranque",
    "compresordeaire",
    "sensorelectrico",
    "tuercassoldadas",
    "conectordeacero",
    "mangueraagranel",
    "deltipocartucho",
    "asientocompleto",
    "placadedesgaste",
    "decauchonitrilo",
    "eslabondecadena",
    "mediaabrazadera",
    "valvuladealivio",
    "panelderadiador",
    "aceroinoxidable",
    "vidriodeventana",
    "herrajesdeacero",
    "cojinetedebiela",
    "placaderespaldo",
    "tornillodeacero",
    "bandadedesgaste",
    "tapadeplastico",
    "aceroyplastico",
    "paralosequipos",
    "resortedeacero",
    "pasadordeacero",
    "filtrodeaceite",
    "sellosdecaucho",
    "adaptadorrecto",
    "tuercasdeacero",
    "soportedeacero",
    "motoniveladora",
    "chavetadeacero",
    "bujeespaciador",
    "arosellocaucho",
    "bujedefriccion",
    "turbocompresor",
    "nucleoradiador",
    "valvuladeacero",
    "clavijadeacero",
    "tapaderadiador",
    "varilladeacero",
    "motorelectrico",
    "planchadeacero",
    "anillodecaucho",
    "cableelectrico",
    "arandeladegoma",
    "conjuntodeguía",
    "mbrazodesello",
    "abrazaderaenu",
    "juntadecaucho",
    "pistondefreno",
    "tuercadeacero",
    "cardandeacero",
    "secadordeaire",
    "dientedebalde",
    "retendeaceite",
    "empaquetadura",
    "pernopasasdor",
    "acopleroscado",
    "deltiporadial",
    "culatademotor",
    "pistondeacero",
    "adaptadorcodo",
    "sellodecaucho",
    "turbocargador",
    "juegodejuntas",
    "plasticoacero",
    "zapatadeoruga",
    "arandelaplana",
    "papelcelulosa",
    "acopledeacero",
    "bombadeaceite",
    "tipobridacat",
    "placasoporte",
    "placadeacero",
    "bujedecaucho",
    "bulóndeacero",
    "juntadepapel",
    "chapadeacero",
    "antifricción",
    "alimentacion",
    "filtrodeaire",
    "ruedadentada",
    "pernopasador",
    "tuboflexible",
    "guardabarros",
    "taponroscado",
    "poleadeacero",
    "anillodegoma",
    "mazodecables",
    "bulondeacero",
    "tipobridasae",
    "tapaplastica",
    "alimentación",
    "tubodecaucho",
    "cauchoyacero",
    "lainagolilla",
    "platodeacero",
    "amortiguador",
    "bridadeacero",
    "pernodeacero",
    "retendelabio",
    "combustible",
    "sellodegoma",
    "hojadecorte",
    "bombadeagua",
    "codoroscado",
    "kitdesellos",
    "bujedeacero",
    "puntadepala",
    "interruptor",
    "tubodeacero",
    "kitdejuntas",
    "cododeacero",
    "resorteagas",
    "anillosello",
    "caterpillar",
    "retencaucho",
    "selloanular",
    "motorescat",
    "juntaoring",
    "provistode",
    "acoplecodo",
    "filtroaire",
    "espaciador",
    "articulado",
    "bulonacero",
    "alternador",
    "arandelasy",
    "calcomania",
    "ventilador",
    "termostato",
    "abrazadera",
    "manometro",
    "deflector",
    "adaptador",
    "escobilla",
    "esparrago",
    "electrico",
    "retenedor",
    "ejecardan",
    "protector",
    "proyectil",
    "engranaje",
    "composite",
    "nitrogeno",
    "aroseguro",
    "sujetador",
    "compresor",
    "actuator",
    "manguito",
    "arosello",
    "plastico",
    "longitud",
    "cargador",
    "aluminio",
    "plástico",
    "manguera",
    "contiene",
    "tornillo",
    "bridasae",
    "cubierta",
    "inyector",
    "elemento",
    "conjunto",
    "pestillo",
    "arandela",
    "cojinete",
    "pasantes",
    "cuchilla",
    "conector",
    "decaucho",
    "celulosa",
    "diámetro",
    "radiador",
    "metalico",
    "juegode",
    "mineros",
    "resorte",
    "interno",
    "montaje",
    "control",
    "espesor",
    "medidas",
    "soporte",
    "amperes",
    "rodillo",
    "tractor",
    "inserto",
    "general",
    "asiento",
    "valvula",
    "fusible",
    "equipos",
    "pasador",
    "chaveta",
    "deacero",
    "apiston",
    "sistema",
    "roscado",
    "dientes",
    "tobera",
    "cables",
    "selloo",
    "grampa",
    "minero",
    "correa",
    "alarma",
    "broche",
    "piston",
    "acople",
    "camion",
    "llanta",
    "cabina",
    "rigido",
    "caudal",
    "varios",
    "flange",
    "ampere",
    "aceite",
    "hembra",
    "puntas",
    "zapata",
    "anillo",
    "tuerca",
    "filtro",
    "caucho",
    "vidrio",
    "radial",
    "cable",
    "aspas",
    "junta",
    "mmthk",
    "motor",
    "chapa",
    "brida",
    "oring",
    "perno",
    "papel",
    "macho",
    "reten",
    "polea",
    "oruga",
    "traba",
    "grupo",
    "bulon",
    "acero",
    "items",
    "juego",
    "laina",
    "entre",
    "sobre",
    "placa",
    "sello",
    "tapon",
    "bomba",
    "laton",
    "cobre",
    "ring",
    "agua",
    "buje",
    "pulg",
    "cono",
    "volt",
    "peso",
    "vela",
    "seal",
    "tubo",
    "para",
    "como",
    "mazo",
    "aire",
    "tipo",
    "codo",
    "tapa",
    "caja",
    "bajo",
    "taza",
    "maza",
    "del",
    "xqc",
    "chd",
    "aro",
    "una",
    "eje",
    "sin",
    "vcc",
    "con",
    "bhd",
    "por",
    "uso",
    "vdc",
    "los",
    "psi",
    "kit",
    "min",
    "thd",
    "las",
    "un",
    "al",
    "de",
    "cm",
    "la",
    "od",
    "mm",
    "id",
    "el",
    "kg",
    "pm",
    "en",
])



def separar_palabras_pegadas(texto):
    """Intenta separar palabras pegadas usando lista de palabras conocidas."""
    # Si el texto ya tiene espacios normales, no hacer nada
    palabras = texto.split()
    resultado = []
    modificado = False

    for palabra in palabras:
        # Ignorar si tiene menos de 8 chars, es número, tiene espacios o es código/modelo
        if (len(palabra) < 8 or
            re.match(r'^[\d\W]+$', palabra) or
            re.match(r'^[A-Z]{1,4}\d+', palabra)):
            resultado.append(palabra)
            continue

        # Intentar segmentar
        segmentada = segmentar(palabra.lower())
        if segmentada and len(segmentada) > 1:
            # Preservar mayúsculas originales de la primera letra
            if palabra[0].isupper():
                segmentada[0] = segmentada[0].capitalize()
            resultado.append(" ".join(segmentada))
            modificado = True
        else:
            resultado.append(palabra)

    return " ".join(resultado), modificado

def segmentar(texto, min_len=2):
    """Segmenta un texto pegado usando el conjunto de palabras del dominio."""
    if not texto: return []
    if texto in PALABRAS_SEPARACION: return [texto]
    for i in range(min_len, len(texto)):
        p1, p2 = texto[:i], texto[i:]
        if p1 in PALABRAS_SEPARACION:
            if not p2: return [p1]
            r = segmentar(p2, min_len)
            if r and r != [p2]: return [p1] + r
            if p2 in PALABRAS_SEPARACION: return [p1, p2]
    return [texto]

PALABRAS_CLAVE = ["KIT", "CONJUNTO", "MANGUERA"]
PALABRAS_CLAVE_FRASE = ["ELEMENTO FILTRANTE"]

CORRECCIONES_ORTOGRAFIA = {
    r'\bpasdor\b': 'pasador',
    r'\bbraxo\b': 'brazo',
    r'\bhidaulico\b': 'hidráulico',
    r'\bhidaulica\b': 'hidráulica',
    r'\bectronico[s]?\b': 'electrónico',
    r'\bectrónico[s]?\b': 'electrónico',
    r'\bdelanero\b': 'delantero',
    r'\bdelanera\b': 'delantera',
    r'\bplastico\b': 'plástico',
    r'\bplastica\b': 'plástica',
    r'\bfundicion\b': 'fundición',
    r'\btransmision\b': 'transmisión',
    r'\bdireccion\b': 'dirección',
    r'\bsuspension\b': 'suspensión',
    r'\binyeccion\b': 'inyección',
    r'\bconexion\b': 'conexión',
    r'\bproteccion\b': 'protección',
    r'\bsujecion\b': 'sujeción',
    r'\brotacion\b': 'rotación',
    r'\bcarcaza\b': 'carcasa',
    r'\bhidraulico\b': 'hidráulico',
    r'\bhidraulica\b': 'hidráulica',
    r'\belectrico\b': 'eléctrico',
    r'\belectrica\b': 'eléctrica',
    r'\bmecanico\b': 'mecánico',
    r'\bneumatico\b': 'neumático',
    r'\bconico\b': 'cónico',
    r'\bconica\b': 'cónica',
    r'\bsintetico\b': 'sintético',
    r'\bsintetica\b': 'sintética',
    r'\bpresion\b': 'presión',
    r'\bvalvula\b': 'válvula',
    r'\bvalvulas\b': 'válvulas',
    r'\bpistons\b': 'pistones',
    r'\bfijacion\b': 'fijación',
    r'\bseparacion\b': 'separación',
    r'\bdistribucion\b': 'distribución',
    r'\bfluoroelastomero\b': 'fluoroelastómero',
}

# ── Funciones ──────────────────────────────────────────────

def limpiar_url(texto):
    return re.sub(r'https?://\S+', '', texto).strip()

def limpiar_codigo_interno(texto):
    """Elimina códigos internos tipo SEAL_EXHAUST_1974834.
    Si el texto es SOLO un código, extrae las palabras descriptivas y las traduce."""
    
    # Detectar si todo el texto es un código tipo PALABRA_PALABRA_NUMERO
    patron_codigo_completo = r'^([A-Z][A-Z_]+)_\d{5,}$'
    match = re.match(patron_codigo_completo, texto.strip())
    
    if match:
        # Extraer partes descriptivas (sin el número final)
        partes = texto.strip().split('_')
        palabras = [p for p in partes if not p.isdigit() and len(p) > 1]
        # Traducir cada parte usando el diccionario
        traducidas = []
        for p in palabras:
            p_lower = p.lower()
            if p_lower in DICCIONARIO_TECNICO:
                traducidas.append(DICCIONARIO_TECNICO[p_lower])
            else:
                traducidas.append(p.capitalize())
        return " ".join(traducidas), True
    
    # Caso normal: eliminar códigos embebidos en el texto
    texto = re.sub(r'\b[A-Z]+_[A-Z]+_\d{5,}\b', '', texto)
    texto = re.sub(r'\b[A-Z_]{3,}_\d{5,}\b', '', texto)
    return re.sub(r'\s+', ' ', texto).strip(), False

def detectar_palabras_clave(texto):
    texto_upper = texto.upper()
    encontradas = []
    # Palabras simples
    for p in PALABRAS_CLAVE:
        if re.search(r'\b' + p + r'\b', texto_upper):
            encontradas.append(p)
    # Frases de dos o más palabras (deben aparecer juntas)
    for frase in PALABRAS_CLAVE_FRASE:
        if frase in texto_upper:
            encontradas.append(frase)
    return " | ".join([f"⚠️ {p}" for p in encontradas]) if encontradas else ""

def corregir_ortografia(texto):
    errores = []
    for patron, correccion in CORRECCIONES_ORTOGRAFIA.items():
        match = re.search(patron, texto, re.IGNORECASE)
        if match:
            errores.append(f"ortografía: {match.group()}→{correccion}")
            texto = re.sub(patron, correccion, texto, flags=re.IGNORECASE)
    return texto, errores

def es_marca_o_modelo(palabra):
    return bool(re.match(r'^(CAT|Caterpillar|CATERPILLAR|SEM|CAT\d+|[A-Z]{1,4}\d+[A-Z]?|[A-Z]\d+[A-Z]\d*)$', palabra))

def es_medida(palabra):
    return bool(re.match(r'^\d+[\.\-,]?\d*\s*(mm|MM|cm|m|psi|PSI|kg|KG|lb|VCC|VCA|rpm|RPM|pulg|\'|\")?$', palabra))

def traducir_token(token):
    """Traduce solo si la palabra está en el diccionario técnico."""
    limpio = token.strip('.,;:()/\'"`°-').lower()
    if limpio in DICCIONARIO_TECNICO:
        return DICCIONARIO_TECNICO[limpio], limpio
    return None, None


def procesar_descripcion(descripcion_original):
    errores_encontrados = []

    # 1. Limpiar URL
    desc = limpiar_url(descripcion_original)
    if desc != descripcion_original:
        errores_encontrados.append("URL eliminada")

    # 2. Limpiar códigos internos
    desc_sin_codigos, fue_solo_codigo = limpiar_codigo_interno(desc)
    if fue_solo_codigo:
        errores_encontrados.append(f"código interno traducido: {desc.strip()}→{desc_sin_codigos}")
    elif desc_sin_codigos != desc:
        errores_encontrados.append("código interno eliminado")
    desc = desc_sin_codigos

    # 3. Separar palabras pegadas
    desc_separada, fue_separada = separar_palabras_pegadas(desc)
    if fue_separada:
        errores_encontrados.append("palabras separadas")
    desc = desc_separada

    # 4. Corregir ortografía
    desc, errores_orto = corregir_ortografia(desc)
    errores_encontrados.extend(errores_orto)

    # 4. Traducir palabra por palabra
    tokens = desc.split()
    tokens_nuevos = []
    for token in tokens:
        if es_marca_o_modelo(token) or es_medida(token):
            tokens_nuevos.append(token)
            continue
        traduccion, original = traducir_token(token)
        if traduccion:
            errores_encontrados.append(f"traducido: {original}→{traduccion}")
            tokens_nuevos.append(traduccion)
        else:
            tokens_nuevos.append(token)

    desc = " ".join(tokens_nuevos)

    # 5. Limpiar espacios y normalizar
    desc = re.sub(r'\s+', ' ', desc).strip()
    if desc:
        desc = desc[0].upper() + desc[1:]

    # 6. Detectar palabras clave
    keywords = detectar_palabras_clave(desc)

    resumen = " | ".join(errores_encontrados) if errores_encontrados else "Sin errores"
    return desc, resumen, keywords


def generar_excel(resultados):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Correcciones"

    header_fill = PatternFill(start_color="1F3A1F", end_color="1F3A1F", fill_type="solid")
    header_font = Font(bold=True, color="B8F542", size=11)
    ok_fill = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")
    error_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
    kw_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    thin = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC')
    )

    encabezados = ["Código", "Descripción Original", "Errores Detectados", "Palabras Clave ⚠️", "Descripción Corregida"]
    for col, titulo in enumerate(encabezados, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    ws.row_dimensions[1].height = 32

    for fila, r in enumerate(resultados, 2):
        for col, val in enumerate([r["codigo"], r["original"], r["errores"], r["keywords"], r["corregida"]], 1):
            cell = ws.cell(row=fila, column=col, value=val)
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        fill = kw_fill if r["keywords"] else (ok_fill if r["errores"] == "Sin errores" else error_fill)
        for col in range(1, 6):
            ws.cell(row=fila, column=col).fill = fill

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 38
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 45

    ws2 = wb.create_sheet("Resumen")
    total = len(resultados)
    sin_errores = sum(1 for r in resultados if r["errores"] == "Sin errores")
    corregidas = sum(1 for r in resultados if r["errores"] not in ["Sin errores", "Sin descripción"])
    con_kw = sum(1 for r in resultados if r["keywords"])
    ws2['A1'] = "RESUMEN DE PROCESAMIENTO"
    ws2['A1'].font = Font(bold=True, size=14)
    for i, (label, val) in enumerate([("Total:", total), ("Sin errores:", sin_errores), ("Corregidas:", corregidas), ("Palabras clave ⚠️:", con_kw)], 3):
        ws2[f'A{i}'] = label
        ws2[f'B{i}'] = val
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 10

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ── UI ─────────────────────────────────────────────────────
st.set_page_config(page_title="Corrector de Descripciones Finning", page_icon="🔧", layout="wide")

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@400;600;700;800&family=Barlow:wght@300;400;500&display=swap');

    .stApp { background: linear-gradient(135deg, #0a1628 0%, #0d2137 50%, #0a1628 100%) !important; }
    .block-container { padding-top: 1.5rem !important; max-width: 1100px !important; }

    .hero-wrap {
        background: linear-gradient(90deg, #0d3b6e 0%, #1a5fa8 60%, #0d3b6e 100%);
        border: 1px solid #1e6ab8;
        border-radius: 12px;
        padding: 2.2rem 2.5rem;
        margin-bottom: 2rem;
        position: relative;
        overflow: hidden;
    }
    .hero-wrap::after {
        content: '⚙';
        position: absolute;
        right: 1.5rem; top: 50%;
        transform: translateY(-50%);
        font-size: 9rem;
        opacity: 0.05;
        line-height: 1;
    }
    .hero-tag {
        background: rgba(255,165,0,0.15);
        border: 1px solid rgba(255,165,0,0.4);
        color: #ffb347;
        font-family: 'Barlow Condensed', sans-serif;
        font-size: 0.75rem;
        letter-spacing: 0.2em;
        text-transform: uppercase;
        padding: 3px 14px;
        border-radius: 3px;
        display: inline-block;
        margin-bottom: 0.7rem;
    }
    .hero-title {
        font-family: 'Barlow Condensed', sans-serif;
        font-size: 3rem;
        font-weight: 800;
        color: #fff;
        line-height: 1.05;
        margin-bottom: 0.4rem;
        letter-spacing: 0.01em;
    }
    .hero-title span { color: #ffb347; }
    .hero-sub {
        font-family: 'Barlow', sans-serif;
        font-weight: 300;
        color: rgba(255,255,255,0.55);
        font-size: 0.92rem;
        letter-spacing: 0.04em;
    }
    .hero-badges { margin-top: 1rem; display: flex; gap: 8px; flex-wrap: wrap; }
    .hbadge {
        background: rgba(255,255,255,0.07);
        border: 1px solid rgba(255,255,255,0.12);
        color: rgba(255,255,255,0.7);
        font-family: 'Barlow', sans-serif;
        font-size: 0.78rem;
        padding: 3px 12px;
        border-radius: 20px;
    }

    h1,h2,h3 { font-family: 'Barlow Condensed', sans-serif !important; color: #fff !important; letter-spacing: 0.02em !important; }
    p, .stMarkdown p { color: rgba(255,255,255,0.8) !important; }

    [data-testid="stFileUploader"] {
        background: rgba(13,59,110,0.25) !important;
        border: 2px dashed #1e6ab8 !important;
        border-radius: 10px !important;
    }

    .stButton > button[kind="primary"] {
        background: linear-gradient(90deg, #e67e00, #ffb347) !important;
        color: #0a1628 !important;
        font-family: 'Barlow Condensed', sans-serif !important;
        font-size: 1.15rem !important;
        font-weight: 800 !important;
        letter-spacing: 0.12em !important;
        text-transform: uppercase !important;
        border: none !important;
        border-radius: 8px !important;
        transition: all 0.2s !important;
    }
    .stButton > button[kind="primary"]:hover {
        box-shadow: 0 6px 20px rgba(255,165,0,0.35) !important;
        transform: translateY(-1px) !important;
    }
    .stDownloadButton > button {
        background: linear-gradient(90deg, #1a5fa8, #2980d4) !important;
        color: #fff !important;
        font-family: 'Barlow Condensed', sans-serif !important;
        font-size: 1.1rem !important;
        font-weight: 700 !important;
        letter-spacing: 0.08em !important;
        text-transform: uppercase !important;
        border: none !important;
        border-radius: 8px !important;
    }

    [data-testid="stMetric"] {
        background: rgba(13,59,110,0.4) !important;
        border: 1px solid #1e6ab8 !important;
        border-radius: 10px !important;
        padding: 1rem !important;
    }
    [data-testid="stMetricValue"] { font-family: 'Barlow Condensed', sans-serif !important; color: #ffb347 !important; font-size: 2.2rem !important; }
    [data-testid="stMetricLabel"] { color: rgba(255,255,255,0.55) !important; }

    .stSuccess { background: rgba(0,100,0,0.2) !important; border-left-color: #2d8a2d !important; }
    .stInfo { background: rgba(13,59,110,0.35) !important; border-left-color: #1e6ab8 !important; }
    hr { border-color: rgba(30,106,184,0.25) !important; }
    .stDataFrame { border-radius: 8px !important; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="hero-wrap">
    <div class="hero-tag">🔧 Finning · Repuestos Maquinaria Pesada</div>
    <div class="hero-title">Corrector de<br><span>Descripciones</span></div>
    <div class="hero-sub">Procesamiento inteligente de descripciones de artículos en español</div>
    <div class="hero-badges">
        <span class="hbadge">✓ Traducción semántica</span>
        <span class="hbadge">✓ Corrección ortográfica</span>
        <span class="hbadge">✓ Normalización</span>
        <span class="hbadge">✓ 100% gratuito</span>
    </div>
</div>
""", unsafe_allow_html=True)

st.divider()

archivo = st.file_uploader("📁 Subí tu Excel (Columna A = Código | Columna B = Descripción)", type=["xlsx", "xls"])

if archivo:
    df = pd.read_excel(archivo, header=0)
    df.columns = [str(c).strip() for c in df.columns]

    c1, c2 = st.columns(2)
    with c1:
        st.success(f"✅ **{archivo.name}** — {len(df)} artículos")
    with c2:
        st.info(f"Columnas: **{' | '.join(df.columns.tolist())}**")

    st.dataframe(df.head(5), use_container_width=True)
    st.divider()

    if st.button("▶ Procesar Descripciones", type="primary", use_container_width=True):
        resultados = []
        total = len(df)
        col_codigo = df.columns[0]
        col_desc = df.columns[1]

        progress_bar = st.progress(0)
        status_text = st.empty()
        log_area = st.empty()
        log_lines = []

        for i, row in df.iterrows():
            codigo = str(row[col_codigo]).strip()
            desc_original = str(row[col_desc]).strip() if pd.notna(row[col_desc]) else ""

            progress_bar.progress((i + 1) / total)
            status_text.markdown(f"⚙️ Procesando **{i+1} de {total}**: `{codigo}`")

            if not desc_original or desc_original == "nan":
                resultados.append({"codigo": codigo, "original": "", "errores": "Sin descripción", "keywords": "", "corregida": ""})
                log_lines.append(f"⬜ [{i+1:03d}] {codigo} → Sin descripción")
            else:
                corregida, errores, keywords = procesar_descripcion(desc_original)
                resultados.append({"codigo": codigo, "original": desc_original, "errores": errores, "keywords": keywords, "corregida": corregida})
                icono = "⚠️" if keywords else ("✅" if errores == "Sin errores" else "✏️")
                log_lines.append(f"{icono} [{i+1:03d}] {codigo} → {corregida[:55]}...")

            log_area.code("\n".join(log_lines[-12:]), language=None)

        progress_bar.progress(1.0)
        status_text.markdown("✅ **¡Procesamiento completado!**")
        st.divider()

        sin_errores = sum(1 for r in resultados if r["errores"] == "Sin errores")
        corregidas = sum(1 for r in resultados if r["errores"] not in ["Sin errores", "Sin descripción"])
        con_kw = sum(1 for r in resultados if r["keywords"])

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total procesados", total)
        c2.metric("✅ Sin errores", sin_errores)
        c3.metric("✏️ Corregidas", corregidas)
        c4.metric("⚠️ Palabras clave", con_kw)

        st.subheader("📋 Resultados")
        df_out = pd.DataFrame(resultados)
        df_out.columns = ["Código", "Descripción Original", "Errores Detectados", "Palabras Clave", "Descripción Corregida"]
        st.dataframe(df_out, use_container_width=True, height=420)

        st.divider()
        excel_buffer = generar_excel(resultados)
        st.download_button(
            label="⬇️ Descargar Excel Corregido",
            data=excel_buffer,
            file_name="descripciones_corregidas.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
